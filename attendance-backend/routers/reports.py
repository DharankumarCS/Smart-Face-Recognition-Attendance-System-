import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
import models
from auth import require_role, get_current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()


def calc_attendance_pct(present, total):
    return round((present / total * 100) if total > 0 else 0, 1)


@router.get("/department")
def department_report(
    db: Session = Depends(get_db),
    current_user=Depends(require_role("HOD")),
):
    total_students = (
        db.query(models.Student)
        .join(models.User)
        .filter(models.User.is_active == True)
        .count()
    )
    total_faculty = (
        db.query(models.User)
        .filter(models.User.role == "Faculty", models.User.is_active == True)
        .count()
    )
    total_subjects = db.query(models.Subject).count()

    all_records = db.query(models.AttendanceRecord).all()
    present = sum(1 for r in all_records if r.status in ("present", "late"))
    total_records = len(all_records) if all_records else 1
    avg_attendance = calc_attendance_pct(present, total_records)

    subject_stats = []
    subjects = db.query(models.Subject).all()
    for sub in subjects:
        sessions = (
            db.query(models.AttendanceSession)
            .filter(models.AttendanceSession.subject_id == sub.id)
            .all()
        )
        session_ids = [s.id for s in sessions]
        if not session_ids:
            subject_stats.append({
                "name": sub.name,
                "code": sub.code,
                "faculty": "",
                "percentage": 0,
            })
            continue
        records = (
            db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.session_id.in_(session_ids))
            .all()
        )
        p = sum(1 for r in records if r.status in ("present", "late"))
        subject_stats.append({
            "name": sub.name,
            "code": sub.code,
            "faculty_id": sub.faculty_id,
            "percentage": calc_attendance_pct(p, len(records)),
        })

    for ss in subject_stats:
        if ss.get("faculty_id"):
            f = db.query(models.User).filter(models.User.id == ss["faculty_id"]).first()
            ss["faculty"] = f.name if f else ""
            del ss["faculty_id"]

    section_stats = []
    for sec in ["A", "B", "C", "D"]:
        students = db.query(models.Student).filter(models.Student.section == sec).all()
        if not students:
            continue
        sids = [s.id for s in students]
        records = (
            db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.student_id.in_(sids))
            .all()
        )
        p = sum(1 for r in records if r.status in ("present", "late"))
        section_stats.append({
            "section": sec,
            "students": len(students),
            "percentage": calc_attendance_pct(p, len(records) if records else len(students)),
        })

    recent_sessions = (
        db.query(models.AttendanceSession)
        .order_by(models.AttendanceSession.created_at.desc())
        .limit(5)
        .all()
    )
    recent = []
    for s in recent_sessions:
        subject = db.query(models.Subject).filter(models.Subject.id == s.subject_id).first()
        faculty = db.query(models.User).filter(models.User.id == s.faculty_id).first()
        recent.append({
            "id": s.id,
            "subject_name": subject.name if subject else "",
            "faculty_name": faculty.name if faculty else "",
            "date": str(s.date),
            "status": s.status,
        })

    return {
        "total_students": total_students,
        "total_faculty": total_faculty,
        "total_subjects": total_subjects,
        "avg_attendance": avg_attendance,
        "subject_stats": subject_stats,
        "section_stats": section_stats,
        "recent_activity": recent,
    }


@router.get("/student/{register_number}")
def student_report(
    register_number: str,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    reg = register_number.upper()
    student = (
        db.query(models.Student)
        .filter(models.Student.register_number == reg)
        .first()
    )
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    if current_user.role == "Student" and current_user.register_number != reg:
        raise HTTPException(status_code=403, detail="Access denied")

    user = db.query(models.User).filter(models.User.id == student.user_id).first()
    subjects = db.query(models.Subject).filter(
        models.Subject.year == student.year,
        models.Subject.section == student.section,
    ).all()

    subject_wise = []
    total_present = 0
    total_classes = 0

    for sub in subjects:
        sessions = (
            db.query(models.AttendanceSession)
            .filter(models.AttendanceSession.subject_id == sub.id)
            .all()
        )
        session_ids = [s.id for s in sessions]
        if not session_ids:
            subject_wise.append({
                "name": sub.name,
                "code": sub.code,
                "present": 0,
                "total": 0,
                "percentage": 0,
            })
            continue
        records = (
            db.query(models.AttendanceRecord)
            .filter(
                models.AttendanceRecord.session_id.in_(session_ids),
                models.AttendanceRecord.student_id == student.id,
            )
            .all()
        )
        present = sum(1 for r in records if r.status in ("present", "late"))
        total = len(records)
        total_present += present
        total_classes += total
        subject_wise.append({
            "name": sub.name,
            "code": sub.code,
            "present": present,
            "total": total,
            "percentage": calc_attendance_pct(present, total),
        })

    overall = calc_attendance_pct(total_present, total_classes)
    needed = 0
    if overall < 75 and total_classes > 0:
        target = int(total_classes * 0.75) + 1
        needed = max(0, target - total_present)

    recent_records = (
        db.query(models.AttendanceRecord)
        .join(models.AttendanceSession)
        .filter(models.AttendanceRecord.student_id == student.id)
        .order_by(models.AttendanceSession.date.desc())
        .limit(10)
        .all()
    )
    recent = []
    for r in recent_records:
        session = (
            db.query(models.AttendanceSession)
            .filter(models.AttendanceSession.id == r.session_id)
            .first()
        )
        subject = (
            db.query(models.Subject)
            .filter(models.Subject.id == session.subject_id)
            .first()
            if session
            else None
        )
        recent.append({
            "date": str(session.date) if session else "",
            "subject": subject.name if subject else "",
            "status": r.status,
        })

    return {
        "name": user.name if user else "",
        "register_number": reg,
        "year": student.year,
        "section": student.section,
        "department": student.department,
        "overall_percentage": overall,
        "total_present": total_present,
        "total_classes": total_classes,
        "below_75": overall < 75,
        "classes_needed": needed,
        "subject_wise": subject_wise,
        "recent": recent,
    }


@router.get("/faculty/{faculty_id}")
def faculty_report(
    faculty_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role == "Faculty" and current_user.id != faculty_id:
        raise HTTPException(status_code=403, detail="Access denied")
    if current_user.role not in ("HOD", "Faculty"):
        raise HTTPException(status_code=403, detail="Access denied")

    faculty = db.query(models.User).filter(models.User.id == faculty_id).first()
    if not faculty:
        raise HTTPException(status_code=404, detail="Faculty not found")

    subjects = db.query(models.Subject).filter(models.Subject.faculty_id == faculty_id).all()
    subject_stats = []
    for sub in subjects:
        sessions = (
            db.query(models.AttendanceSession)
            .filter(
                models.AttendanceSession.subject_id == sub.id,
                models.AttendanceSession.faculty_id == faculty_id,
            )
            .all()
        )
        session_ids = [s.id for s in sessions]
        avg_pct = 0
        if session_ids:
            records = (
                db.query(models.AttendanceRecord)
                .filter(models.AttendanceRecord.session_id.in_(session_ids))
                .all()
            )
            p = sum(1 for r in records if r.status in ("present", "late"))
            avg_pct = calc_attendance_pct(p, len(records))

        subject_stats.append({
            "name": sub.name,
            "code": sub.code,
            "classes_taken": len(sessions),
            "avg_attendance": avg_pct,
        })

    recent_sessions = (
        db.query(models.AttendanceSession)
        .filter(models.AttendanceSession.faculty_id == faculty_id)
        .order_by(models.AttendanceSession.created_at.desc())
        .limit(10)
        .all()
    )
    recent = []
    for s in recent_sessions:
        subject = db.query(models.Subject).filter(models.Subject.id == s.subject_id).first()
        records = (
            db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.session_id == s.id)
            .all()
        )
        p = sum(1 for r in records if r.status in ("present", "late"))
        recent.append({
            "id": s.id,
            "subject": subject.name if subject else "",
            "date": str(s.date),
            "status": s.status,
            "present": p,
            "total": len(records),
        })

    student_count = 0
    if subjects:
        sub = subjects[0]
        student_count = (
            db.query(models.Student)
            .filter(
                models.Student.year == sub.year,
                models.Student.section == sub.section,
            )
            .count()
        )

    return {
        "name": faculty.name,
        "subjects_count": len(subjects),
        "students_count": student_count,
        "subject_stats": subject_stats,
        "recent_sessions": recent,
    }


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    current_user=Depends(require_role("HOD")),
):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

    ws1 = wb.active
    ws1.title = "Student Attendance"
    headers1 = ["Register No", "Name", "Year", "Section", "Overall %"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    students = db.query(models.Student).all()
    row = 2
    for s in students:
        user = db.query(models.User).filter(models.User.id == s.user_id).first()
        records = (
            db.query(models.AttendanceRecord)
            .filter(models.AttendanceRecord.student_id == s.id)
            .all()
        )
        p = sum(1 for r in records if r.status in ("present", "late"))
        pct = calc_attendance_pct(p, len(records) if records else 1)
        ws1.cell(row=row, column=1, value=s.register_number)
        ws1.cell(row=row, column=2, value=user.name if user else "")
        ws1.cell(row=row, column=3, value=s.year)
        ws1.cell(row=row, column=4, value=s.section)
        ws1.cell(row=row, column=5, value=pct)
        row += 1

    ws2 = wb.create_sheet("Section Summary")
    headers2 = ["Section", "Students", "Attendance %"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, sec in enumerate(["A", "B", "C", "D"], 2):
        count = db.query(models.Student).filter(models.Student.section == sec).count()
        ws2.cell(row=i, column=1, value=sec)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=0)

    ws3 = wb.create_sheet("Faculty Report")
    headers3 = ["Name", "Email", "Subjects", "Department"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    faculty_list = (
        db.query(models.User)
        .filter(models.User.role == "Faculty")
        .all()
    )
    for i, f in enumerate(faculty_list, 2):
        sub_count = db.query(models.Subject).filter(models.Subject.faculty_id == f.id).count()
        ws3.cell(row=i, column=1, value=f.name)
        ws3.cell(row=i, column=2, value=f.email)
        ws3.cell(row=i, column=3, value=sub_count)
        ws3.cell(row=i, column=4, value=f.department)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"},
    )
